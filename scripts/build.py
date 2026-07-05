#!/usr/bin/env python3
"""
Prakrati Leads CRM — end-to-end build pipeline.

Regenerates ../index.html from the raw source data:
  source-data/Real Estate Data.xlsx   (Bitrix24-style kanban export)
  scripts/img_ocr_leads.py            (OCR transcriptions of the 5 photo lists)
  source-data/*.jpg                   (photos, base64-embedded for in-app verify)
  scripts/template.html               (HTML shell with __DATA__ placeholder)

Steps: parse Excel -> merge photo OCR leads -> phone country classification
(libphonenumber) -> embed images -> inject JSON into template -> write index.html

Run:  python3 scripts/build.py
Deps: pip install openpyxl phonenumbers
"""
import os, re, json, base64, collections
import openpyxl
import phonenumbers as pn
from phonenumbers import is_valid_number, parse as pnparse, region_code_for_number
from img_ocr_leads import IMG_LEADS          # list of photo-OCR lead dicts

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SRC  = os.path.join(ROOT, "source-data")
XLSX = os.path.join(SRC, "Real Estate Data.xlsx")
IMAGES = ["0f446bd0-8e73-46bd-a82a-03fc5232fa28.jpg","3733d235-0d95-46e1-a62d-597cc5ca6ed6.jpg",
          "3bfe2974-f503-4bb0-9bd3-dc2ba7597f20.jpg","8380efe1-9609-4e66-a228-fe4b9750f5ce.jpg",
          "bfd53b94-ddbf-411b-be01-a4bae9670d96.jpg"]

# ---------------------------------------------------------------- Excel parse
# The sheet is a vertical kanban export. Data lives in columns A(1),C(3),E(5),F(6).
# Each lead is a block of labelled rows: <name> / 'Responsible person' / <resp>
# / 'First Name' / <first> / 'Phone No.' / <phone> / '+ Activity' / ...
STAGE = {1:"Assigned", 3:"List B", 5:"List C", 6:"List D"}
LABELS = {"Responsible person","First Name","Phone No.","+ Activity","Assigned"}

def clean(v):
    if v is None: return None
    if isinstance(v,str):
        s=v.replace("\xa0"," ").strip(); return s or None
    if isinstance(v,float) and v.is_integer(): return str(int(v))
    return str(v)

def parse_excel():
    ws = openpyxl.load_workbook(XLSX, data_only=True)["Sheet1"]
    leads=[]
    for c in (1,3,5,6):
        cells=[clean(ws.cell(r,c).value) for r in range(1,ws.max_row+1)]
        cells=[x for x in cells if x is not None]
        n=len(cells)
        for i in range(n-1):
            if cells[i+1]=="Responsible person" and cells[i] not in LABELS:
                rec={"name":cells[i],"responsible":None,"first_name":None,"phone_raw":None,"stage":STAGE[c]}
                j=i+1
                while j<n:
                    lbl=cells[j]
                    if lbl=="Responsible person" and j+1<n: rec["responsible"]=cells[j+1]; j+=2; continue
                    if lbl=="First Name" and j+1<n: rec["first_name"]=cells[j+1]; j+=2; continue
                    if lbl=="Phone No." and j+1<n: rec["phone_raw"]=cells[j+1]; j+=2; continue
                    if lbl=="+ Activity": break
                    j+=1
                leads.append(rec)
    return leads

def extract_numbers(raw):
    if not raw: return []
    parts=re.split(r"[\/,;&]| or | OR ", str(raw))
    out=[]; seen=set()
    for p in parts:
        d=re.sub(r"\D","",p)
        if len(d)>=7 and d not in seen:
            seen.add(d); out.append(d)
    return out

def dedupe_excel(rows):
    merged={}
    for l in rows:
        nums=extract_numbers(l["phone_raw"])
        key=(str(l["name"]).lower().strip(), nums[0] if nums else "")
        if key in merged:
            m=merged[key]
            if l["stage"] not in m["stages"]: m["stages"].append(l["stage"])
            if l["responsible"] and l["responsible"] not in m["responsibles"]: m["responsibles"].append(l["responsible"])
        else:
            merged[key]={"name":l["name"],"first_name":l["first_name"],
                         "responsibles":[l["responsible"]] if l["responsible"] else [],
                         "phone_raw":l["phone_raw"],"phones":nums,"stages":[l["stage"]]}
    return list(merged.values())

# ------------------------------------------------------------ classification
FLAG={c:"".join(chr(0x1F1E6+ord(x)-65) for x in c) for c in
      "AE RU NG EG DE AU JP ES PY IN PK GB US FR KZ SA QA KW OM BH TR UA BY IT CA PH LB SY IQ JO MA DZ TN ID CN ZA NL CH PL".split()}
CNAME={"AE":"UAE","RU":"Russia","NG":"Nigeria","EG":"Egypt","DE":"Germany","AU":"Australia","JP":"Japan","ES":"Spain",
       "PY":"Paraguay","IN":"India","PK":"Pakistan","GB":"UK","US":"USA","FR":"France","KZ":"Kazakhstan","SA":"Saudi Arabia",
       "QA":"Qatar","KW":"Kuwait","OM":"Oman","BH":"Bahrain","TR":"Turkey","UA":"Ukraine"}
GTMAP={"australia":"AU","paraguay":"PY","spain":"ES","uae":"AE","egypt":"EG","japan":"JP","germany":"DE"}
CC={"AE":"971","RU":"7","NG":"234","IN":"91","EG":"20","DE":"49","AU":"61","JP":"81","ES":"34","PY":"595"}
SLAV=("ov","ova","ev","eva","in","ina","sky","skaya","enko","vich","uk","yan")
IND=set("gurav shinde nighot nomulwar lengare mujumdar patel kumar singh sharma reddy rao gupta nair menon iyer pillai desai joshi mehta shah verma yadav khan das bose roy naidu patil jadhav pawar deshmukh kulkarni bhosale more shetty swapnil amol amit sachin abhijeet sanjyoti rahul rohit vijay ajay sanjay anil sunil ramesh suresh mahesh".split())

def name_origin(name):
    toks=re.split(r"\s+",(name or "").lower())
    for t in toks:
        if t in IND: return "IN"
    for t in toks:
        for s in SLAV:
            if t.endswith(s) and len(t)>4: return "RU"
    return None

def valid(e):
    try:
        n=pnparse(e,None); return is_valid_number(n), region_code_for_number(n)
    except: return False,None

def classify(d, name="", gt=None):
    d=re.sub(r"\D","",d or "")
    if not d: return "",None,"none"
    if gt and gt.lower() in GTMAP:
        reg=GTMAP[gt.lower()]
        try:
            n=pnparse(d,reg)
            if is_valid_number(n): return pn.format_number(n,pn.PhoneNumberFormat.E164),reg,"high"
        except: pass
        return "+"+CC.get(reg,"")+d.lstrip("0"),reg,"low"
    opts=[]
    if d.startswith("971") and len(d)>=11: opts.append(("+"+d,"AE","high"))
    if len(d)==9 and d[0]=="5": opts.append(("+971"+d,"AE","high"))
    if len(d)==10 and d[:2]=="05": opts.append(("+971"+d[1:],"AE","high"))
    if len(d)==11 and d[0]=="0": opts.append(("+234"+d[1:],"NG","high"))
    if len(d)==10 and d[0] in "678": opts.append(("+91"+d,"IN","medium"))
    if len(d)==10 and d[0]=="9":
        o=name_origin(name)
        if o=="IN": opts.append(("+91"+d,"IN","medium"))
        elif o=="RU": opts.append(("+7"+d,"RU","medium"))
        else: opts+=[("+7"+d,"RU","low"),("+91"+d,"IN","low")]
    if len(d)==11 and d[0] in "78": opts.append(("+7"+d[1:],"RU","high"))
    for cc,reg in [("20","EG"),("49","DE"),("61","AU"),("81","JP"),("34","ES"),("595","PY"),("44","GB"),
                   ("91","IN"),("92","PK"),("90","TR"),("966","SA"),("974","QA"),("965","KW"),("968","OM"),("973","BH"),("7","RU")]:
        if d.startswith(cc) and len(d)>=10: opts.append(("+"+d,reg,"high"))
    opts+=[("+"+d,None,"low"),("+971"+d,"AE","low")]
    for e,reg,conf in opts:
        ok,r=valid(e)
        if ok: return e,(r or reg),conf
    if len(d)==9 and d[0]=="5": return "+971"+d,"AE","medium"
    return "+"+d,None,"low"

# ------------------------------------------------------------------- build
def main():
    excel=dedupe_excel(parse_excel())
    print(f"excel leads: {len(excel)}   photo leads: {len(IMG_LEADS)}")
    records=[]; i=0; ctry=collections.Counter(); conf=collections.Counter()

    for f in excel:
        i+=1
        res=[classify(p,f["name"]) for p in f["phones"]]
        prim=res[0] if res else ("",None,"none")
        ctry[prim[1]]+=1; conf[prim[2]]+=1
        records.append({"id":i,"name":f["name"],"first_name":f.get("first_name") or f["name"],
            "phone_primary":prim[0],"phone_all":" , ".join(x[0] for x in res),
            "phone_raw":str(f["phone_raw"]) if f["phone_raw"] is not None else "",
            "country_iso":prim[1] or "","country":CNAME.get(prim[1],""),"flag":FLAG.get(prim[1],""),
            "phone_confidence":prim[2],
            "responsible":" / ".join(f["responsibles"]) if f["responsibles"] else "",
            "stage":" , ".join(f["stages"]),"status":"New","last_contacted":"","source":"Excel","notes":"","source_image":""})
    for f in IMG_LEADS:
        i+=1
        res=[classify(p,f["name"],f.get("country","")) for p in f["phones"]]
        prim=res[0] if res else ("",None,"none")
        ctry[prim[1]]+=1
        c="verify" if prim[0] else "none"; conf[c]+=1
        records.append({"id":i,"name":f["name"],"first_name":f["name"],
            "phone_primary":prim[0],"phone_all":" , ".join(x[0] for x in res),
            "phone_raw":f["phone_raw"],
            "country_iso":prim[1] or "","country":CNAME.get(prim[1],""),"flag":FLAG.get(prim[1],""),
            "phone_confidence":c,"responsible":"","stage":"Photo list","status":"New","last_contacted":"",
            "source":"Photo (OCR)","notes":("code:"+f["extra"]) if f.get("extra") else "","source_image":f.get("source_image","")})

    print("total:",len(records),"| by country:",ctry.most_common(),"| by confidence:",conf.most_common())

    imgs={fn:"data:image/jpeg;base64,"+base64.b64encode(open(os.path.join(SRC,fn),"rb").read()).decode() for fn in IMAGES}
    data=json.dumps({"records":records,"images":imgs},ensure_ascii=False)

    tmpl=open(os.path.join(HERE,"template.html")).read()
    html=tmpl.replace("__DATA__",data)
    out=os.path.join(ROOT,"index.html")
    open(out,"w").write(html)
    print("wrote",out,f"({len(html)/1e6:.2f} MB)")

if __name__=="__main__":
    main()
